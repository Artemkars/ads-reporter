"""
report/excel_writer.py

Формирование Excel-отчёта (.xlsx) по данным ClientReport.

Структура выходного файла:
  - Один лист "Отчёт"
  - Заголовок: клиент + период + курс + НДС
  - Шапка таблицы (строка 4)
  - Строки по кампаниям
  - Итоговая строка (жирная, выделена цветом)
  - Нижний колонтитул с датой формирования

Имя файла: AMK_report_2026-W32.xlsx
"""
import logging
import os
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

from .calculator import ClientReport, ReportRow

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Цветовая схема
# ---------------------------------------------------------------------------
COLOR_HEADER_BG = "1F3864"    # тёмно-синий — шапка таблицы
COLOR_HEADER_FG = "FFFFFF"    # белый текст в шапке
COLOR_TOTAL_BG  = "D6E4F0"    # светло-голубой — итоговая строка
COLOR_TITLE_FG  = "1F3864"    # тёмно-синий — заголовок отчёта
COLOR_META_FG   = "4472C4"    # синий Meta — метка источника
COLOR_ACTIVE    = "27AE60"    # зелёный — активная кампания
COLOR_PAUSED    = "E74C3C"    # красный — остановленная
COLOR_ODD_ROW   = "EBF5FB"    # голубоватый фон для чётных строк
COLOR_BORDER    = "BDC3C7"    # серый — линии таблицы

# ---------------------------------------------------------------------------
# Колонки таблицы
# ---------------------------------------------------------------------------
COLUMNS = [
    ("Кампания",          50, "left"),
    ("Статус",            14, "center"),
    ("Расход, $",         14, "right"),
    ("Расход, ₸ (с НДС)", 18, "right"),
    ("Результаты",        10, "center"),
    ("Цена результата, ₸", 16, "right"),
]


def write_excel_report(
    reports: list[ClientReport], output_dir: str = "output"
) -> str:
    """Генерирует Excel-отчет с группировкой (outlineLevel) и цветовой заливкой."""
    if not reports:
        return ""
        
    client_name = reports[0].client_name
    date_label = reports[0].date_label.replace(" ", "").replace("-", "_")

    out_dir = Path(output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    filepath = str(out_dir / f"Отчет_{client_name}_{date_label}.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = f"Отчет {client_name}"

    current_row = 1
    data_start_row = 1

    for client_report in reports:
        current_row = _write_header(ws, client_report, current_row)
        current_row = _write_table_header(ws, current_row)
    
        block_data_start = current_row
        if current_row == data_start_row or data_start_row == 1:
            data_start_row = block_data_start
            
        for i, row in enumerate(client_report.rows):
            current_row = _write_campaign_row(ws, row, current_row, is_even=(i % 2 == 1))
    
        _write_total_row(ws, client_report.total, current_row)
        current_row += 1
    
        current_row += 1
        _write_footer(ws, client_report, current_row, block_data_start)
        
        current_row += 3 # Spacing between blocks

    # --- Общая сводка (Grand Summary) ---
    if len(reports) > 1:
        total_spend_usd = sum(r.total.spend_usd for r in reports)
        total_spend_kzt = sum(r.total.spend_kzt for r in reports)
        total_leads = 0
        
        for r in reports:
            if r.original_campaigns:
                for camp in r.original_campaigns:
                    if camp.level == 'campaign' and camp.is_lead_campaign:
                        total_leads += camp.results
                        
        avg_cost = round(total_spend_kzt / total_leads, 2) if total_leads > 0 else 0.0
        
        # Заголовок сводки
        title_cell = ws.cell(row=current_row, column=1)
        title_cell.value = "ОБЩАЯ СВОДКА (Meta + Google)"
        title_cell.font = Font(name="Calibri", size=12, bold=True, color=COLOR_TITLE_FG)
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        _merge(ws, current_row, 1, current_row, len(COLUMNS))
        current_row += 1
        
        # Шапка сводки
        headers = ["Общий расход $", "Общий расход ₸", "Всего лидов", "Ср. цена лида"]
        thin = Side(style="thin", color=COLOR_BORDER)
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for i, h in enumerate(headers, start=1):
            cell = ws.cell(row=current_row, column=i)
            cell.value = h
            cell.font = Font(name="Calibri", size=10, bold=True, color=COLOR_HEADER_FG)
            cell.fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        current_row += 1
        
        # Значения сводки
        values = [
            (round(total_spend_usd, 2), "#,##0.00"), 
            (round(total_spend_kzt, 2), "#,##0 ₸"), 
            (total_leads, "0"), 
            (avg_cost, "#,##0 ₸")
        ]
        for i, (v, num_fmt) in enumerate(values, start=1):
            cell = ws.cell(row=current_row, column=i)
            cell.value = v
            cell.number_format = num_fmt
            cell.font = Font(name="Calibri", size=10, bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        
        current_row += 2

    # --- Ширина колонок -----------------------------------------------------
    for col_idx, (_, width, _) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # --- Заморозить шапку (строка с названиями колонок) ---------------------
    ws.freeze_panes = ws.cell(row=data_start_row, column=1)

    wb.save(filepath)
    logger.info(
        "[%s] Excel-отчёт сохранён: %s (%d источников)",
        client_name.upper(), filepath, len(reports)
    )

    return filepath


# ---------------------------------------------------------------------------
# Вспомогательные функции
# ---------------------------------------------------------------------------

def _write_header(ws, client_report: ClientReport, row: int) -> int:
    """Пишет шапку документа (название, период, курс, НДС)."""
    num_cols = len(COLUMNS)

    # Строка 1: Название клиента + источник
    title_cell = ws.cell(row=row, column=1)
    title_cell.value = f"Рекламный отчёт — {client_report.client_name}"
    title_cell.font = Font(
        name="Calibri", size=16, bold=True, color=COLOR_TITLE_FG
    )
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 28
    _merge(ws, row, 1, row, num_cols)
    row += 1

    # Строка 2: Период
    period_cell = ws.cell(row=row, column=1)
    period_cell.value = f"Период: {client_report.date_label}  |  Источник: Meta Ads"
    period_cell.font = Font(name="Calibri", size=10, italic=True, color=COLOR_META_FG)
    _merge(ws, row, 1, row, num_cols)
    row += 1

    # Строка 3: Курс и НДС
    params_cell = ws.cell(row=row, column=1)
    params_cell.value = (
        f"Курс USD/₸: {client_report.rate_usd_kzt:,.2f}  |  НДС + АК: {client_report.vat_pct}%"
    )
    params_cell.font = Font(name="Calibri", size=10, color="7F8C8D")
    _merge(ws, row, 1, row, num_cols)
    row += 1

    # Пустая строка-разделитель
    row += 1

    return row


def _write_table_header(ws, row: int) -> int:
    """Рисует шапку таблицы."""
    thin = Side(style="thin", color=COLOR_BORDER)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, (col_name, _, align) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = col_name
        cell.font = Font(name="Calibri", size=10, bold=True, color=COLOR_HEADER_FG)
        cell.fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
        cell.alignment = Alignment(
            horizontal=align, vertical="center", wrap_text=True
        )
        cell.border = border

    ws.row_dimensions[row].height = 32
    return row + 1


def _write_campaign_row(
    ws, row_data: ReportRow, row: int, is_even: bool
) -> int:
    """Пишет одну строку кампании."""
    thin = Side(style="thin", color=COLOR_BORDER)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    if row_data.level == "campaign":
        ws.row_dimensions[row].outlineLevel = 0
        bg_color = "E8F0FE" # Light blue
    elif row_data.level == "adset":
        ws.row_dimensions[row].outlineLevel = 1
        bg_color = "F3F2F1" # Light gray
    elif row_data.level == "ad":
        ws.row_dimensions[row].outlineLevel = 2
        bg_color = "FFFFFF" # White
    else:
        bg_color = COLOR_ODD_ROW if is_even else "FFFFFF"

    values = [
        row_data.name,
        row_data.status,
        row_data.spend_usd,
        row_data.spend_kzt,
        row_data.results,
        row_data.cost_per_result,
    ]
    aligns = ["left", "center", "right", "right", "center", "right"]
    formats = [None, None, '#,##0.00', '#,##0.00 ₸', '0', '#,##0.00 ₸']

    for col_idx, (value, align, fmt) in enumerate(zip(values, aligns, formats), start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = value
        cell.font = Font(name="Calibri", size=10)
        cell.fill = PatternFill("solid", fgColor=bg_color)
        cell.alignment = Alignment(horizontal=align, vertical="center")
        cell.border = border
        if fmt:
            cell.number_format = fmt

    # Цвет статуса
    status_cell = ws.cell(row=row, column=2)
    if row_data.status == "Активна":
        status_cell.font = Font(name="Calibri", size=10, color=COLOR_ACTIVE, bold=True)
    elif row_data.status in ("Остановлена", "Отключена"):
        status_cell.font = Font(name="Calibri", size=10, color=COLOR_PAUSED)

    ws.row_dimensions[row].height = 20
    return row + 1


def _write_total_row(ws, total: ReportRow, row: int) -> None:
    """Рисует итоговую строку."""
    medium = Side(style="medium", color=COLOR_HEADER_BG)
    border = Border(left=medium, right=medium, top=medium, bottom=medium)

    values = [
        total.name,
        "",
        total.spend_usd,
        total.spend_kzt,
        total.results,
        total.cost_per_result,
    ]
    aligns = ["left", "center", "right", "right", "center", "right"]
    formats = [None, None, '#,##0.00', '#,##0.00 ₸', '0', '#,##0.00 ₸']

    for col_idx, (value, align, fmt) in enumerate(zip(values, aligns, formats), start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = value
        cell.font = Font(
            name="Calibri", size=10, bold=True, color=COLOR_TITLE_FG
        )
        cell.fill = PatternFill("solid", fgColor=COLOR_TOTAL_BG)
        cell.alignment = Alignment(horizontal=align, vertical="center")
        cell.border = border
        if fmt:
            cell.number_format = fmt

    ws.row_dimensions[row].height = 22


def _write_footer(
    ws, client_report: ClientReport, row: int, data_start_row: int
) -> None:
    """Пишет нижний колонтитул с мета-информацией."""
    generated_at = datetime.now().strftime("%d.%m.%Y %H:%M")
    num_cols = len(COLUMNS)

    footer_cell = ws.cell(row=row, column=1)
    footer_cell.value = (
        f"Сформировано: {generated_at}  |  "
        f"Строк данных: {len(client_report.rows)}  |  "
        f"Расход итого: {client_report.total.spend_kzt:,.0f} ₸  |  "
        f"Результатов итого: {client_report.total.results}"
    )
    footer_cell.font = Font(name="Calibri", size=9, italic=True, color="95A5A6")
    _merge(ws, row, 1, row, num_cols)


def _merge(ws, r1: int, c1: int, r2: int, c2: int) -> None:
    """Безопасное объединение ячеек."""
    ws.merge_cells(
        start_row=r1, start_column=c1,
        end_row=r2, end_column=c2
    )
